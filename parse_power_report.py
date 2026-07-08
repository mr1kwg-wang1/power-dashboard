"""
성신양회 단양공장 전력월보 자동 파싱 스크립트
사용법: python parse_power_report.py <전력월보_YYYY-MM.xlsx>
동일 폴더의 data.json 에 해당 월 데이터를 추가/갱신한다.

전제: "전력월보_YYYY-MM.xlsx" 형식의 파일명, "전력월보"/"전력요금계산서" 시트 구조가
2026-06 원본 템플릿과 동일해야 함 (셀 좌표가 바뀌면 COORDS 만 수정하면 됨).
"""
import sys, json, re
from pathlib import Path
from openpyxl import load_workbook

COORDS = {
    "OR":  ("AD13", "AE13"),
    "RM":  ("AD19", "AE19"),
    "COM": ("AD24", "AE24"),
    "KL":  ("AD37", "AE37"),
    "CM":  ("AD44", "AE44"),
    "usage_subtotal": ("AD48", "AE48"),
    "total": ("AD54", "AE54"),
}
RATE_COORDS = {
    "usage_fee_won": "G10",
    "base_fee_won": "G20",
    "total_fee_won": "G31",
    "peak_mid_kw": "D19",
    "peak_max_kw": "E19",
    "unit_price_won_per_kwh": "F28",
    "period": "B3",
}

ROW_MAP = {
    "OR":  13,
    "RM":  19,
    "COM": 24,
    "KL":  37,
    "CM":  44,
    "usage_subtotal": 48,
    "total": 54,
}

def find_total_col(ws, header_row=5, search_row=6):
    """헤더 행에서 '계' 라벨이 있는 열을 찾아 (사용량열, 원단위열) 컬럼 인덱스를 반환."""
    from openpyxl.utils import get_column_letter
    for col in range(1, 60):
        v = ws.cell(row=header_row, column=col).value
        if v is not None and str(v).strip() == "계":
            # 다음 non-empty 라벨 열 전까지가 이 '계' 블록. 보통 usage_col=col, unit_col=col+1
            usage_col = col
            unit_col = col + 1
            return usage_col, unit_col
    raise ValueError("헤더 행에서 '계' 열을 찾을 수 없습니다.")

def parse_file(path: Path):
    wb = load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    ws1 = wb[sheet_names[0]]
    ws2 = wb[sheet_names[1]] if len(sheet_names) > 1 else None

    m = re.search(r"(\d{4})[-_](\d{2})", path.stem)
    if not m:
        raise ValueError(f"파일명에서 연-월을 찾을 수 없습니다: {path.name} (예: 전력월보_2026-06.xlsx)")
    year, month = m.group(1), m.group(2)

    usage_col, unit_col = find_total_col(ws1)

    processes = {}
    for key, r in ROW_MAP.items():
        usage = ws1.cell(row=r, column=usage_col).value
        unit = ws1.cell(row=r, column=unit_col).value
        processes[key] = {
            "usage_kwh": round(usage, 0) if isinstance(usage, (int, float)) else None,
            "unit_kwh_t": round(unit, 2) if isinstance(unit, (int, float)) else None,
        }

    rate = {}
    if ws2 is not None:
        for key, cell in RATE_COORDS.items():
            v = ws2[cell].value
            if key == "period":
                rate[key] = v
            elif isinstance(v, (int, float)):
                rate[key] = round(v, 2)
            else:
                rate[key] = v

    return {
        "month": f"{year}-{month}",
        "source": "file",
        "file_name": path.name,
        "processes": {
            "OR": processes["OR"],
            "RM": processes["RM"],
            "COM": processes["COM"],
            "KL": processes["KL"],
            "CM": processes["CM"],
        },
        "usage_subtotal": processes["usage_subtotal"],
        "total": processes["total"],
        "rate": rate,
    }

def main():
    if len(sys.argv) < 2:
        print("사용법: python parse_power_report.py <전력월보_YYYY-MM.xlsx>")
        sys.exit(1)

    src = Path(sys.argv[1])
    entry = parse_file(src)

    data_path = Path(__file__).parent / "data.json"
    if data_path.exists():
        data = json.loads(data_path.read_text(encoding="utf-8"))
    else:
        data = {"months": []}

    # 기존에 실무자가 남긴 note(수기 메모)는 재파싱해도 보존
    existing = next((m for m in data["months"] if m["month"] == entry["month"]), None)
    if existing and existing.get("note"):
        entry["note"] = existing["note"]

    data["months"] = [m for m in data["months"] if m["month"] != entry["month"]]
    data["months"].append(entry)
    data["months"].sort(key=lambda m: m["month"])

    data_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] {entry['month']} 데이터를 data.json 에 반영했습니다. (총 {len(data['months'])}개월 누적)")

if __name__ == "__main__":
    main()
